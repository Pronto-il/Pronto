#!/usr/bin/env python3
"""Converts the human-readable classification workbook into the two machine-readable
artefacts the backend consumes, deterministically and without ever writing to the workbook.

    python backend/tools/classification_dataset/build_dataset.py \
        --workbook ~/Downloads/pronto_classification_experiment_5000_v2.xlsx

Outputs (both regenerated from scratch, both checked in):

  backend/src/main/resources/ai/profession-taxonomy.json
      The 50 professions x 250 subcategories that make up the CLASSIFICATION label space,
      plus each profession's DISPATCH mapping onto the seven rows of the production
      `categories` table. Loaded at startup by ai.taxonomy.ProfessionTaxonomy.

  backend/src/test/resources/ai-eval/classification-dataset-v2.jsonl
      One line per workbook row, carrying the original `ID` so any result can always be
      traced back to the spreadsheet, plus its frozen dev/validation/holdout split.

WHY A GENERATOR RATHER THAN HAND-MAINTAINED FILES. The workbook is the human-readable
source of truth and it will keep moving; 250 subcategory codes and 5,000 split assignments
transcribed by hand would drift from it silently. Re-running this script is how the code
and the spreadsheet are kept provably in step -- it fails loudly when the workbook stops
matching the taxonomy authored here, which is the check that matters.

DETERMINISM IS THE WHOLE POINT. Nothing here reads the clock, the filesystem ordering or a
random seed. Splits are a pure function of (dataset ID, SPLIT_SALT), so the same row lands
in the same split on every machine and on every run -- which is what makes "Prompt V2 beat
V1 on the validation set" a comparison rather than a coincidence. Changing SPLIT_SALT
reshuffles every split and invalidates every previously reported number; it is a constant,
not a knob.
"""

import argparse
import collections
import hashlib
import json
import pathlib
import sys

# --------------------------------------------------------------------------------------
# Split policy
# --------------------------------------------------------------------------------------

#: Fixed for the life of the dataset. See the module docstring -- changing it silently
#: invalidates every accuracy number ever reported against a split.
SPLIT_SALT = "pronto-classification-v2"

#: 70 / 15 / 15, applied WITHIN each (profession, subcategory) group rather than across the
#: dataset as a whole. Every group holds exactly 20 rows, so this is exact -- 14/3/3 -- and
#: every one of the 250 groups is represented in all three splits. A global hash split would
#: have left some groups absent from validation entirely, which would make per-group accuracy
#: on those splits undefined rather than merely noisy.
DEV_PER_GROUP = 14
VALIDATION_PER_GROUP = 3
HOLDOUT_PER_GROUP = 3
GROUP_SIZE = DEV_PER_GROUP + VALIDATION_PER_GROUP + HOLDOUT_PER_GROUP

SPLIT_DEV = "dev"
SPLIT_VALIDATION = "validation"
SPLIT_HOLDOUT = "holdout"

# --------------------------------------------------------------------------------------
# The classification taxonomy
# --------------------------------------------------------------------------------------
#
# Hebrew is the workbook's key and stays the customer-facing label; the codes are what the
# model returns and what the evaluation reports on. Ordering matters: the five subcategory
# codes line up positionally with the five workbook rows for that profession, and the script
# asserts that below rather than trusting it.

# fmt: off
PROFESSIONS = [
    ("אינסטלטור", "PLUMBER", [
        ("פתיחת סתימות", "CLOGGED_DRAIN"),
        ("נזילה מברז/חיבור", "FAUCET_OR_CONNECTION_LEAK"),
        ("ניאגרה ואסלה", "TOILET_OR_CISTERN"),
        ("לחץ/זרימת מים", "WATER_PRESSURE_OR_FLOW"),
        ("פיצוץ/נזילה משמעותית בצנרת", "BURST_PIPE_OR_MAJOR_LEAK"),
    ]),
    ("חשמלאי", "ELECTRICIAN", [
        ("קצר/קפיצת חשמל", "SHORT_CIRCUIT_OR_BREAKER_TRIP"),
        ("שקע/מתג תקול", "FAULTY_OUTLET_OR_SWITCH"),
        ("לוח חשמל", "DISTRIBUTION_BOARD"),
        ("תאורה", "LIGHTING"),
        ("נקודת חשמל/התקנה", "NEW_POINT_OR_INSTALLATION"),
    ]),
    ("טכנאי מזגנים", "AC_TECHNICIAN", [
        ("לא מקרר", "NOT_COOLING"),
        ("לא מחמם", "NOT_HEATING"),
        ("נזילת מים", "WATER_LEAK"),
        ("רעש/רעידות", "NOISE_OR_VIBRATION"),
        ("לא נדלק/נכבה", "NOT_TURNING_ON_OR_SHUTS_DOWN"),
    ]),
    ("טכנאי דודים", "BOILER_TECHNICIAN", [
        ("אין מים חמים", "NO_HOT_WATER"),
        ("דוד לא נדלק/תקלה חשמלית", "NOT_TURNING_ON_OR_ELECTRICAL"),
        ("נזילה מהדוד", "BOILER_LEAK"),
        ("בעיה בקולט שמש", "SOLAR_COLLECTOR"),
        ("זרימת מים חמים חלשה", "WEAK_HOT_WATER_FLOW"),
    ]),
    ("מנעולן", "LOCKSMITH", [
        ("דלת נעולה/לא נפתחת", "LOCKED_OUT"),
        ("החלפת צילינדר", "CYLINDER_REPLACEMENT"),
        ("מפתח תקוע/שבור", "KEY_STUCK_OR_BROKEN"),
        ("מנגנון נעילה תקול", "LOCKING_MECHANISM_FAULT"),
        ("פריצה/פתיחת דלת", "EMERGENCY_DOOR_OPENING"),
    ]),
    ("הנדימן", "HANDYMAN", [
        ("תלייה וקידוח", "HANGING_AND_DRILLING"),
        ("הרכבת רהיטים", "FURNITURE_ASSEMBLY"),
        ("תיקוני פרזול/ידיות", "HARDWARE_AND_HANDLES"),
        ("איטום סיליקון", "SILICONE_SEALING"),
        ("התקנות ביתיות", "SMALL_HOME_INSTALLATIONS"),
    ]),
    ("טכנאי מקררים", "REFRIGERATOR_TECHNICIAN", [
        ("מקרר לא מקרר", "NOT_COOLING"),
        ("מקפיא לא מקפיא", "FREEZER_NOT_FREEZING"),
        ("נזילת מים/קרח", "WATER_OR_ICE_LEAK"),
        ("רעש חריג", "ABNORMAL_NOISE"),
        ("לא נדלק/עובד רצוף", "NOT_TURNING_ON_OR_RUNS_CONSTANTLY"),
    ]),
    ("טכנאי מכונות כביסה", "WASHING_MACHINE_TECHNICIAN", [
        ("לא סוחטת/לא מסתובבת", "NOT_SPINNING"),
        ("לא מרוקנת מים", "NOT_DRAINING"),
        ("נזילת מים", "WATER_LEAK"),
        ("לא נדלקת/תוכנית תקועה", "NOT_STARTING_OR_CYCLE_STUCK"),
        ("רעש/רעידות", "NOISE_OR_VIBRATION"),
    ]),
    ("טכנאי מדיחי כלים", "DISHWASHER_TECHNICIAN", [
        ("לא מנקה", "NOT_CLEANING"),
        ("לא מרוקן מים", "NOT_DRAINING"),
        ("נזילה", "LEAK"),
        ("לא נדלק/לא מתחיל", "NOT_STARTING"),
        ("לא מחמם/מייבש", "NOT_HEATING_OR_DRYING"),
    ]),
    ("טכנאי תנורים וכיריים", "OVEN_AND_COOKTOP_TECHNICIAN", [
        ("תנור לא מחמם", "OVEN_NOT_HEATING"),
        ("חימום לא אחיד", "UNEVEN_HEATING"),
        ("כיריים לא נדלקות", "COOKTOP_NOT_IGNITING"),
        ("הצתה/ניצוץ", "IGNITION_OR_SPARK"),
        ("דלת/כפתורים/תצוגה", "DOOR_KNOBS_OR_DISPLAY"),
    ]),
    ("טכנאי מייבשי כביסה", "DRYER_TECHNICIAN", [
        ("לא מייבש", "NOT_DRYING"),
        ("לא מסתובב", "NOT_TUMBLING"),
        ("לא נדלק/נכבה", "NOT_TURNING_ON_OR_SHUTS_DOWN"),
        ("רעש", "NOISE"),
        ("ניקוז/עיבוי", "DRAINAGE_OR_CONDENSATION"),
    ]),
    ("שיפוצניק", "RENOVATION_CONTRACTOR", [
        ("שיפוץ חדר/דירה קטן", "SMALL_ROOM_OR_APARTMENT_RENOVATION"),
        ("שבירת/בניית קיר", "WALL_DEMOLITION_OR_BUILD"),
        ("תיקוני טיח ובטון", "PLASTER_AND_CONCRETE_REPAIR"),
        ("שיפוץ אמבטיה", "BATHROOM_RENOVATION"),
        ("תיקוני גמר", "FINISHING_WORK"),
    ]),
    ("צבעי", "PAINTER", [
        ("צביעת חדר/דירה", "ROOM_OR_APARTMENT_PAINTING"),
        ("תיקוני קיר לפני צבע", "WALL_PREP_BEFORE_PAINTING"),
        ("כתמי רטיבות אחרי תיקון", "MOISTURE_STAINS_AFTER_REPAIR"),
        ("צביעת תקרה", "CEILING_PAINTING"),
        ("דלתות/משקופים", "DOORS_AND_FRAMES"),
    ]),
    ("איש גבס", "DRYWALL_INSTALLER", [
        ("תיקון קיר גבס", "DRYWALL_REPAIR"),
        ("בניית מחיצה", "PARTITION_BUILD"),
        ("הנמכת תקרה", "DROPPED_CEILING"),
        ("נישה/מדפים", "NICHE_OR_SHELVING"),
        ("נזקי מים בגבס", "WATER_DAMAGE_TO_DRYWALL"),
    ]),
    ("רצף", "TILER", [
        ("החלפת אריחים שבורים", "BROKEN_TILE_REPLACEMENT"),
        ("ריצוף חדר/שטח", "FLOOR_TILING"),
        ("אריחים רופפים", "LOOSE_TILES"),
        ("רובה", "GROUT"),
        ("חיפוי קיר", "WALL_CLADDING"),
    ]),
    ("קבלן איטום", "WATERPROOFING_CONTRACTOR", [
        ("רטיבות מהגג", "ROOF_MOISTURE"),
        ("איטום מרפסת", "BALCONY_WATERPROOFING"),
        ("איטום קיר חיצוני", "EXTERIOR_WALL_WATERPROOFING"),
        ("איטום חלון/פתח", "WINDOW_OR_OPENING_SEALING"),
        ("איטום חדר רחצה", "BATHROOM_WATERPROOFING"),
    ]),
    ("גגן ורעפים", "ROOFER", [
        ("רעפים שבורים/זזים", "BROKEN_OR_SHIFTED_TILES"),
        ("נזילה מגג רעפים", "TILED_ROOF_LEAK"),
        ("ניקוי/תחזוקת גג", "ROOF_CLEANING_AND_MAINTENANCE"),
        ("איטום חיבורים בגג", "ROOF_JOINT_SEALING"),
        ("תיקון תשתית גג", "ROOF_STRUCTURE_REPAIR"),
    ]),
    ("זגג", "GLAZIER", [
        ("זכוכית שבורה", "BROKEN_GLASS"),
        ("מראה", "MIRROR"),
        ("דלת זכוכית", "GLASS_DOOR"),
        ("מדף/פלטת זכוכית", "GLASS_SHELF_OR_PANEL"),
        ("זכוכית בידודית", "INSULATED_GLAZING"),
    ]),
    ("איש אלומיניום", "ALUMINIUM_INSTALLER", [
        ("חלון אלומיניום תקול", "FAULTY_ALUMINIUM_WINDOW"),
        ("דלת אלומיניום", "ALUMINIUM_DOOR"),
        ("אטימה בחלון", "WINDOW_SEALING"),
        ("מסילה/גלגלים", "TRACK_OR_ROLLERS"),
        ("התקנת חלון/סגירה", "WINDOW_OR_ENCLOSURE_INSTALLATION"),
    ]),
    ("טכנאי תריסים", "SHUTTER_TECHNICIAN", [
        ("תריס חשמלי לא עולה/יורד", "ELECTRIC_SHUTTER_STUCK"),
        ("רצועה/מנגנון ידני", "STRAP_OR_MANUAL_MECHANISM"),
        ("שלבים שבורים", "BROKEN_SLATS"),
        ("מנוע תריס", "SHUTTER_MOTOR"),
        ("כיוון/מסילות", "ALIGNMENT_OR_TRACKS"),
    ]),
    ("מתקין רשתות נגד יתושים", "INSECT_SCREEN_INSTALLER", [
        ("רשת קרועה", "TORN_SCREEN"),
        ("רשת יצאה מהמסילה", "SCREEN_OFF_TRACK"),
        ("התקנת רשת חדשה", "NEW_SCREEN_INSTALLATION"),
        ("רשת לדלת/ויטרינה", "DOOR_OR_PATIO_SCREEN"),
        ("רשת עמידה לחיות", "PET_RESISTANT_SCREEN"),
    ]),
    ("נגר", "CARPENTER", [
        ("ארונות ורהיטי עץ", "CABINETS_AND_WOODEN_FURNITURE"),
        ("רהיט שבור", "BROKEN_FURNITURE"),
        ("נגרות לפי מידה", "CUSTOM_CARPENTRY"),
        ("תיקון/חידוש עץ", "WOOD_REPAIR_OR_RESTORATION"),
        ("מדפים/פתרונות עץ", "SHELVES_AND_WOOD_SOLUTIONS"),
    ]),
    ("מתקין מטבחים", "KITCHEN_INSTALLER", [
        ("דלתות מטבח וצירים", "KITCHEN_DOORS_AND_HINGES"),
        ("מגירות מטבח", "KITCHEN_DRAWERS"),
        ("ארונות מטבח", "KITCHEN_CABINETS"),
        ("התאמה למכשירי חשמל", "APPLIANCE_FIT"),
        ("חזיתות/שדרוג מטבח", "FRONTS_OR_KITCHEN_UPGRADE"),
    ]),
    ("איש שיש", "STONE_AND_COUNTERTOP_FITTER", [
        ("שבר/סדק בשיש", "CHIP_OR_CRACK"),
        ("חידוש/ליטוש", "RESTORATION_OR_POLISHING"),
        ("חיתוך לכיור/כיריים", "SINK_OR_COOKTOP_CUTOUT"),
        ("הדבקה וחיבורים", "BONDING_AND_JOINTS"),
        ("התקנת/החלפת משטח", "COUNTERTOP_INSTALLATION"),
    ]),
    ("מסגר", "METALWORKER", [
        ("שער/דלת מתכת", "METAL_GATE_OR_DOOR"),
        ("מעקה", "RAILING"),
        ("סורגים", "WINDOW_BARS"),
        ("ריתוך ותיקון", "WELDING_AND_REPAIR"),
        ("קונסטרוקציה קטנה", "SMALL_STRUCTURES"),
    ]),
    ("מתקין דלתות", "DOOR_TECHNICIAN", [
        ("דלת משפשפת/לא מיושרת", "DOOR_SCRAPING_OR_MISALIGNED"),
        ("צירים", "HINGES"),
        ("ידית/פרזול", "HANDLE_OR_HARDWARE"),
        ("משקוף", "DOOR_FRAME"),
        ("החלפת/התקנת דלת", "DOOR_REPLACEMENT_OR_INSTALLATION"),
    ]),
    ("מתקין פרקטים", "LAMINATE_FLOORING_INSTALLER", [
        ("פרקט נפוח/רטוב", "SWOLLEN_OR_WET_FLOORING"),
        ("לוחות פגומים", "DAMAGED_PLANKS"),
        ("פרקט חורק/זז", "SQUEAKING_OR_SHIFTING"),
        ("פנלים", "SKIRTING_BOARDS"),
        ("התקנת פרקט", "FLOORING_INSTALLATION"),
    ]),
    ("מתקין טפטים", "WALLPAPER_INSTALLER", [
        ("טפט מתקלף", "PEELING_WALLPAPER"),
        ("בועות/קמטים", "BUBBLES_OR_WRINKLES"),
        ("הסרת טפט", "WALLPAPER_REMOVAL"),
        ("התקנת טפט", "WALLPAPER_INSTALLATION"),
        ("תיקון נקודתי", "SPOT_REPAIR"),
    ]),
    ("ניקיון דירות", "HOME_CLEANING", [
        ("ניקיון חד פעמי", "ONE_TIME_CLEANING"),
        ("ניקיון אחרי שיפוץ", "POST_RENOVATION_CLEANING"),
        ("ניקיון לפני/אחרי מעבר", "MOVE_IN_OR_OUT_CLEANING"),
        ("מטבח/אמבטיה עמוק", "DEEP_KITCHEN_OR_BATHROOM"),
        ("ניקוי חלונות", "WINDOW_CLEANING"),
    ]),
    ("ניקוי ספות ושטיחים", "UPHOLSTERY_AND_CARPET_CLEANING", [
        ("ניקוי ספה", "SOFA_CLEANING"),
        ("ניקוי שטיח", "CARPET_CLEANING"),
        ("ריח מחיות", "PET_ODOUR"),
        ("כתמים קשים", "TOUGH_STAINS"),
        ("מזרנים/כורסאות", "MATTRESSES_AND_ARMCHAIRS"),
    ]),
    ("פוליש וליטוש רצפות", "FLOOR_POLISHING", [
        ("פוליש לרצפה", "FLOOR_POLISH"),
        ("ליטוש אבן/שיש", "STONE_OR_MARBLE_POLISHING"),
        ("הסרת כתמים", "STAIN_REMOVAL"),
        ("קריסטל", "CRYSTALLISATION"),
        ("ניקוי עמוק", "DEEP_CLEANING"),
    ]),
    ("מדביר", "PEST_CONTROL", [
        ("תיקנים", "COCKROACHES"),
        ("נמלים", "ANTS"),
        ("יתושים/זבובים", "MOSQUITOES_AND_FLIES"),
        ("פרעושים/קרציות", "FLEAS_AND_TICKS"),
        ("מכרסמים", "RODENTS"),
    ]),
    ("גנן", "GARDENER", [
        ("תחזוקת גינה", "GARDEN_MAINTENANCE"),
        ("השקיה", "IRRIGATION"),
        ("שתילה והקמה", "PLANTING_AND_SETUP"),
        ("דשא טבעי", "NATURAL_LAWN"),
        ("גיזום שיחים", "SHRUB_PRUNING"),
    ]),
    ("גוזם עצים", "TREE_TRIMMER", [
        ("גיזום עץ", "TREE_PRUNING"),
        ("ענפים יבשים", "DEAD_BRANCHES"),
        ("עץ קרוב למבנה", "TREE_NEAR_STRUCTURE"),
        ("פינוי גזם/עץ מת", "GREEN_WASTE_REMOVAL"),
        ("עיצוב/הקטנת נוף", "CROWN_REDUCTION"),
    ]),
    ("מתקין דשא סינתטי", "ARTIFICIAL_GRASS_INSTALLER", [
        ("התקנת דשא", "GRASS_INSTALLATION"),
        ("דשא התרומם/זז", "LIFTED_OR_SHIFTED_GRASS"),
        ("תפרים וחיבורים", "SEAMS_AND_JOINTS"),
        ("ניקוז ותשתית", "DRAINAGE_AND_BASE"),
        ("ניקוי/רענון", "CLEANING_OR_REFRESH"),
    ]),
    ("טכנאי שערים חשמליים", "ELECTRIC_GATE_TECHNICIAN", [
        ("שער לא נפתח/נסגר", "GATE_NOT_OPENING_OR_CLOSING"),
        ("מנוע שער", "GATE_MOTOR"),
        ("שלט/מקלט", "REMOTE_OR_RECEIVER"),
        ("חיישנים", "SENSORS"),
        ("מסילה/גלגלים", "TRACK_OR_ROLLERS"),
    ]),
    ("טכנאי אינטרקום", "INTERCOM_TECHNICIAN", [
        ("אין צליל/שיחה", "NO_SOUND_OR_CALL"),
        ("פתיחת דלת", "DOOR_RELEASE"),
        ("מסך/וידאו", "SCREEN_OR_VIDEO"),
        ("קודן", "KEYPAD"),
        ("התקנה/שדרוג", "INSTALLATION_OR_UPGRADE"),
    ]),
    ("טכנאי מצלמות אבטחה", "SECURITY_CAMERA_TECHNICIAN", [
        ("מצלמה לא עובדת", "CAMERA_NOT_WORKING"),
        ("אין תמונה/תמונה גרועה", "NO_OR_POOR_IMAGE"),
        ("הקלטה", "RECORDING"),
        ("גישה מהטלפון", "REMOTE_ACCESS"),
        ("התקנת מצלמות", "CAMERA_INSTALLATION"),
    ]),
    ("טכנאי אזעקות", "ALARM_TECHNICIAN", [
        ("אזעקה מופעלת לשווא", "FALSE_ALARMS"),
        ("גלאי תקול", "FAULTY_DETECTOR"),
        ("קוד/לוח מקשים", "CODE_OR_KEYPAD"),
        ("סוללה/תקלה מערכתית", "BATTERY_OR_SYSTEM_FAULT"),
        ("התקנה/הרחבה", "INSTALLATION_OR_EXPANSION"),
    ]),
    ("טכנאי רשתות ביתיות", "HOME_NETWORK_TECHNICIAN", [
        ("אין אינטרנט בנקודה", "NO_INTERNET_AT_POINT"),
        ("Wi‑Fi חלש", "WEAK_WIFI"),
        ("נקודת רשת חדשה", "NEW_NETWORK_POINT"),
        ("ארון תקשורת", "COMMS_CABINET"),
        ("ראוטר/מגדיל טווח", "ROUTER_OR_EXTENDER"),
    ]),
    ("מתקין טלוויזיות וקולנוע ביתי", "TV_AND_HOME_CINEMA_INSTALLER", [
        ("תליית טלוויזיה", "TV_MOUNTING"),
        ("חיבור ציוד", "EQUIPMENT_CONNECTION"),
        ("קולנוע ביתי", "HOME_CINEMA"),
        ("הסתרת כבלים", "CABLE_CONCEALMENT"),
        ("כיוון/התקנה מחדש", "REALIGNMENT_OR_REINSTALL"),
    ]),
    ("טכנאי מחשבים", "COMPUTER_TECHNICIAN", [
        ("מחשב לא נדלק", "COMPUTER_NOT_TURNING_ON"),
        ("מחשב איטי/נתקע", "SLOW_OR_FREEZING"),
        ("אינטרנט/רשת", "INTERNET_OR_NETWORK"),
        ("חומרה/שדרוג", "HARDWARE_OR_UPGRADE"),
        ("מערכת הפעלה", "OPERATING_SYSTEM"),
    ]),
    ("טכנאי משאבות מים", "WATER_PUMP_TECHNICIAN", [
        ("משאבה לא עובדת", "PUMP_NOT_WORKING"),
        ("לחץ מים חלש", "LOW_WATER_PRESSURE"),
        ("רעש/רעידות", "NOISE_OR_VIBRATION"),
        ("נזילה", "LEAK"),
        ("עובדת רצוף", "RUNS_CONSTANTLY"),
    ]),
    ("ביובית", "SEWAGE_TANKER", [
        ("סתימת ביוב ראשי", "MAIN_SEWER_BLOCKAGE"),
        ("שורשים בביוב", "ROOTS_IN_SEWER"),
        ("שאיבת הצפה", "FLOOD_PUMPING"),
        ("שטיפת קו", "LINE_JETTING"),
        ("צילום קו ביוב", "SEWER_CAMERA_INSPECTION"),
    ]),
    ("מאתר נזילות", "LEAK_DETECTION", [
        ("רטיבות בקיר", "WALL_MOISTURE"),
        ("רטיבות בתקרה", "CEILING_MOISTURE"),
        ("חשבון מים חריג", "ABNORMAL_WATER_BILL"),
        ("נזילה מתחת לרצפה", "UNDER_FLOOR_LEAK"),
        ("איתור מקור חדירה", "INGRESS_SOURCE_TRACING"),
    ]),
    ("טכנאי ברי מים", "WATER_BAR_TECHNICIAN", [
        ("אין מים/זרימה חלשה", "NO_WATER_OR_WEAK_FLOW"),
        ("מים לא קרים", "WATER_NOT_COLD"),
        ("מים לא חמים", "WATER_NOT_HOT"),
        ("נזילה", "LEAK"),
        ("פילטר/טעם/התראה", "FILTER_TASTE_OR_ALERT"),
    ]),
    ("טכנאי טוחני אשפה", "GARBAGE_DISPOSAL_TECHNICIAN", [
        ("לא נדלק", "NOT_TURNING_ON"),
        ("תקוע", "JAMMED"),
        ("נזילה", "LEAK"),
        ("רעש חריג", "ABNORMAL_NOISE"),
        ("התקנה/החלפה", "INSTALLATION_OR_REPLACEMENT"),
    ]),
    ("טכנאי גז", "GAS_TECHNICIAN", [
        ("כיריים גז לא נדלקות", "GAS_COOKTOP_NOT_IGNITING"),
        ("ריח/חשד לדליפת גז", "SUSPECTED_GAS_LEAK"),
        ("התקנת כיריים", "COOKTOP_INSTALLATION"),
        ("נקודת גז", "GAS_POINT"),
        ("וסת/זרימת גז", "REGULATOR_OR_GAS_FLOW"),
    ]),
    ("מתקין סוככים ופרגולות", "AWNING_AND_PERGOLA_INSTALLER", [
        ("סוכך לא נפתח/נסגר", "AWNING_STUCK"),
        ("בד קרוע/בלוי", "TORN_OR_WORN_FABRIC"),
        ("מנוע/שלט", "MOTOR_OR_REMOTE"),
        ("פרגולה תקולה", "FAULTY_PERGOLA"),
        ("התקנה/שדרוג", "INSTALLATION_OR_UPGRADE"),
    ]),
    ("מתקין מקלחונים", "SHOWER_ENCLOSURE_INSTALLER", [
        ("דלת לא נסגרת", "DOOR_NOT_CLOSING"),
        ("נזילה", "LEAK"),
        ("צירים/גלגלים", "HINGES_OR_ROLLERS"),
        ("זכוכית/אטמים", "GLASS_OR_SEALS"),
        ("התקנה", "INSTALLATION"),
    ]),
]
# fmt: on

# --------------------------------------------------------------------------------------
# The dispatch mapping -- the SECOND, separate concern
# --------------------------------------------------------------------------------------
#
# A profession maps to a production category if and only if that category's authored scope in
# ai.catalog.CategoryRoutingProfiles ALREADY covers the work. The rule is deliberately
# mechanical: every entry below is justified by a sentence that exists in that file today, so
# this map records an existing product boundary rather than inventing a new one.
#
# Absence from this map is a first-class, correct answer -- it means Pronto classified the
# request correctly and cannot currently dispatch it, which is the UNSUPPORTED_PROFESSION
# flow that already exists. It is NEVER a reason to route the customer to the nearest
# available trade instead.
DISPATCH = {
    # plumbing -- "the building's water supply, drainage and sanitary systems, plus the
    # domestic water heater (boiler) unit itself".
    "PLUMBER": "plumbing",
    # "No hot water, or hot water only from the water-heater unit failing". Pronto has no
    # separate boiler category by design; V29's plumbing_boiler_replace sub-service is the
    # same decision expressed in the database.
    "BOILER_TECHNICIAN": "plumbing",
    # "Blocked drains, sinks, showers, toilets or main sewage lines" -- main-line blockages
    # are explicitly in scope today, and demoting them to unsupported would regress a case
    # that currently books.
    "SEWAGE_TANKER": "plumbing",
    # "Sewage smell, damp patch or water stain traced to a pipe run", plus the plumbing/
    # painting overlap rule: an unexplained damp patch goes to plumbing FIRST so the source
    # is found. That is leak detection's job description.
    "LEAK_DETECTION": "plumbing",
    # "Low or fluctuating water pressure across the home"; pressure reducers are named as a
    # typical plumbing component.
    "WATER_PUMP_TECHNICIAN": "plumbing",
    "ELECTRICIAN": "electrical",
    "AC_TECHNICIAN": "ac_hvac",
    # appliance_repair -- "free-standing and built-in domestic appliances as self-contained
    # machines", which names washing machines, dryers, dishwashers, fridges, freezers, ovens
    # cooktops and microwaves outright.
    "REFRIGERATOR_TECHNICIAN": "appliance_repair",
    "WASHING_MACHINE_TECHNICIAN": "appliance_repair",
    "DISHWASHER_TECHNICIAN": "appliance_repair",
    "OVEN_AND_COOKTOP_TECHNICIAN": "appliance_repair",
    "DRYER_TECHNICIAN": "appliance_repair",
    "GARBAGE_DISPOSAL_TECHNICIAN": "appliance_repair",
    "WATER_BAR_TECHNICIAN": "appliance_repair",
    "LOCKSMITH": "locksmith",
    "PAINTER": "painting",
    "HANDYMAN": "general_handyman",
    # "Adjusting a door or cabinet that rubs, sags or will not close properly" is listed
    # under general_handyman, and the locksmith<->handyman overlap rule is written around
    # exactly this door-leaf-versus-lock boundary.
    "DOOR_TECHNICIAN": "general_handyman",
}

#: The seven rows of the production `categories` table (V10 as amended by V31). Listed so the
#: script can reject a typo in DISPATCH rather than emitting a dangling category code that
#: would only fail at application startup.
PRODUCTION_CATEGORY_CODES = {
    "plumbing", "electrical", "ac_hvac", "appliance_repair",
    "locksmith", "painting", "general_handyman",
}

TAXONOMY_VERSION = "profession-taxonomy-v1"
DATASET_VERSION = "pronto_classification_experiment_5000_v2"


def split_for(dataset_id, group_ids):
    """The frozen split for one row.

    Rank inside the row's own (profession, subcategory) group by a salted hash of the
    dataset ID, then take the first 14 as dev, the next 3 as validation and the last 3 as
    holdout. Depends only on the ID set and SPLIT_SALT, so it is identical on every run and
    every machine, and it is unaffected by prompt or taxonomy changes.
    """
    ranked = sorted(group_ids, key=lambda i: hashlib.sha256(f"{SPLIT_SALT}:{i}".encode()).hexdigest())
    position = ranked.index(dataset_id)
    if position < DEV_PER_GROUP:
        return SPLIT_DEV
    if position < DEV_PER_GROUP + VALIDATION_PER_GROUP:
        return SPLIT_VALIDATION
    return SPLIT_HOLDOUT


def main():
    parser = argparse.ArgumentParser(description=__doc__)
    parser.add_argument("--workbook", required=True, type=pathlib.Path)
    parser.add_argument("--repo", type=pathlib.Path, default=pathlib.Path(__file__).resolve().parents[3])
    args = parser.parse_args()

    try:
        import openpyxl
    except ImportError:
        sys.exit("openpyxl is required:  python -m pip install openpyxl")

    if not args.workbook.exists():
        sys.exit(f"workbook not found: {args.workbook}")

    # read_only + data_only: this script must never be capable of writing to the workbook.
    workbook = openpyxl.load_workbook(args.workbook, data_only=True, read_only=True)

    # ---- validate the authored taxonomy against the workbook -----------------------------
    taxonomy_rows = [r for r in workbook["Taxonomy"].iter_rows(values_only=True)][1:]
    workbook_taxonomy = collections.OrderedDict()
    for rank, category, subcategory, seed, _examples in taxonomy_rows:
        if category is None:
            continue
        workbook_taxonomy.setdefault(category, []).append((subcategory, seed))

    authored_names = [name for name, _code, _subs in PROFESSIONS]
    if authored_names != list(workbook_taxonomy):
        missing = set(workbook_taxonomy) - set(authored_names)
        extra = set(authored_names) - set(workbook_taxonomy)
        sys.exit(f"taxonomy drift.\n  in workbook, not authored: {sorted(missing)}\n"
                 f"  authored, not in workbook: {sorted(extra)}")

    professions = []
    for rank, (name_he, code, subs) in enumerate(PROFESSIONS, start=1):
        workbook_subs = workbook_taxonomy[name_he]
        if [s for s, _ in workbook_subs] != [s for s, _ in subs]:
            sys.exit(f"subcategory drift under {name_he} ({code}):\n"
                     f"  workbook: {[s for s, _ in workbook_subs]}\n"
                     f"  authored: {[s for s, _ in subs]}")
        dispatch = DISPATCH.get(code)
        if dispatch is not None and dispatch not in PRODUCTION_CATEGORY_CODES:
            sys.exit(f"{code} dispatches to '{dispatch}', which is not a production category code")
        professions.append({
            "code": code,
            "nameHe": name_he,
            "priorityRank": rank,
            "dispatchCategoryCode": dispatch,
            "subcategories": [
                {"code": sub_code, "nameHe": sub_he, "seedSymptom": seed}
                for (sub_he, sub_code), (_, seed) in zip(subs, workbook_subs)
            ],
        })

    duplicate_codes = [c for c, n in collections.Counter(p["code"] for p in professions).items() if n > 1]
    if duplicate_codes:
        sys.exit(f"duplicate profession codes: {duplicate_codes}")

    taxonomy = {
        "taxonomyVersion": TAXONOMY_VERSION,
        "sourceWorkbook": args.workbook.name,
        "generatedBy": "backend/tools/classification_dataset/build_dataset.py",
        "notes": (
            "CLASSIFICATION label space (professions x subcategories) and, separately, each "
            "profession's DISPATCH mapping onto the production categories table. A null "
            "dispatchCategoryCode means Pronto classifies the request correctly and cannot "
            "currently dispatch it -- the UNSUPPORTED_PROFESSION flow -- and is never a reason "
            "to substitute a different profession. Regenerate with the script above; do not "
            "hand-edit."
        ),
        "professions": professions,
    }

    taxonomy_path = args.repo / "backend/src/main/resources/ai/profession-taxonomy.json"
    taxonomy_path.parent.mkdir(parents=True, exist_ok=True)
    taxonomy_path.write_text(
        json.dumps(taxonomy, ensure_ascii=False, indent=2) + "\n", encoding="utf-8")

    # ---- convert the 5,000 labelled rows -------------------------------------------------
    dataset_rows = [r for r in workbook["Dataset"].iter_rows(values_only=True)]
    header = list(dataset_rows[0])
    column = {name: index for index, name in enumerate(header)}
    profession_by_name_he = {p["nameHe"]: p for p in professions}

    records = []
    for row in dataset_rows[1:]:
        category_he = row[column["Expected Category"]]
        subcategory_he = row[column["Expected Subcategory"]]
        profession = profession_by_name_he.get(category_he)
        if profession is None:
            sys.exit(f"row {row[column['ID']]}: unknown Expected Category {category_he!r}")
        subcategory = next(
            (s for s in profession["subcategories"] if s["nameHe"] == subcategory_he), None)
        if subcategory is None:
            sys.exit(f"row {row[column['ID']]}: {subcategory_he!r} is not a subcategory of {category_he!r}")

        records.append({
            "id": int(row[column["ID"]]),
            "description": row[column["User Description"]],
            "expectedProfession": profession["code"],
            "expectedSubcategory": subcategory["code"],
            "expectedIntent": row[column["Expected Intent"]],
            "expectedUrgency": row[column["Expected Urgency"]],
            "expectedNeedsClarification": row[column["Needs Clarification"]] == "YES",
            "expectedDispatchCategory": profession["dispatchCategoryCode"],
            "descriptionStyle": row[column["Description Style"]],
            "edgeCase": row[column["Edge Case"]] == "כן",
            "evalType": row[column["Eval Type"]],
        })

    groups = collections.defaultdict(list)
    for record in records:
        groups[(record["expectedProfession"], record["expectedSubcategory"])].append(record["id"])
    for key, ids in groups.items():
        if len(ids) != GROUP_SIZE:
            sys.exit(f"group {key} holds {len(ids)} rows; the {DEV_PER_GROUP}/"
                     f"{VALIDATION_PER_GROUP}/{HOLDOUT_PER_GROUP} split assumes exactly {GROUP_SIZE}")

    for record in records:
        record["split"] = split_for(
            record["id"], groups[(record["expectedProfession"], record["expectedSubcategory"])])

    records.sort(key=lambda r: r["id"])
    dataset_path = args.repo / "backend/src/test/resources/ai-eval/classification-dataset-v2.jsonl"
    dataset_path.parent.mkdir(parents=True, exist_ok=True)
    with dataset_path.open("w", encoding="utf-8", newline="\n") as handle:
        for record in records:
            handle.write(json.dumps(record, ensure_ascii=False, sort_keys=True) + "\n")

    # ---- manifest: how a reported number is traced back to its inputs --------------------
    digest = hashlib.sha256(dataset_path.read_bytes()).hexdigest()
    split_counts = collections.Counter(r["split"] for r in records)
    manifest = {
        "datasetVersion": DATASET_VERSION,
        "taxonomyVersion": TAXONOMY_VERSION,
        "sourceWorkbook": args.workbook.name,
        "splitSalt": SPLIT_SALT,
        "rows": len(records),
        "professions": len(professions),
        "subcategories": sum(len(p["subcategories"]) for p in professions),
        "dispatchable": sum(1 for p in professions if p["dispatchCategoryCode"]),
        "splitCounts": dict(sorted(split_counts.items())),
        "datasetSha256": digest,
    }
    manifest_path = args.repo / "backend/src/test/resources/ai-eval/classification-dataset-v2.manifest.json"
    manifest_path.write_text(
        json.dumps(manifest, ensure_ascii=False, indent=2) + "\n", encoding="utf-8")

    print(f"taxonomy  -> {taxonomy_path}")
    print(f"dataset   -> {dataset_path}")
    print(f"manifest  -> {manifest_path}")
    print(json.dumps(manifest, ensure_ascii=False, indent=2))


if __name__ == "__main__":
    main()
